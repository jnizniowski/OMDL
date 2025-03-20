import toml
import random
import time
import openpyxl
import sys
import json
import os
import pickle
import re

from typing import Dict, Any, Tuple, List
from pathlib import Path
from queue import Queue
from threading import Thread, Event
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from collections import deque
from datetime import datetime
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

##### PROJECT INFO
PROJECT_NAME = "OMDL (Oh My DataLayer)"
PROJECT_VERSION = "1.0"
PROJECT_AUTHOR = "Jakub Niżniowski"
PROJECT_URL = "https://github.com/jnizniowski/OMDL"
PROJECT_DESCRIPTION = "DataLayer events scraping & validation tool"

PROJECT_LICENSE = "MIT License"
PROJECT_COPYRIGHT = "Copyright (c) 2025 Jakub Niżniowski"

PROJECT_HEADER = f"""
{'='*60}
{PROJECT_NAME} v{PROJECT_VERSION}
{PROJECT_DESCRIPTION}
Created by {PROJECT_AUTHOR}
Documentation: {PROJECT_URL}
{'='*60}
"""

allowed_validation_types = {'<int>', '<float>', '<str>'}

##### CLASSES

class LogCollector:
    """Collect log messages with timestamps for debugging"""
    def __init__(self, console_levels=("INFO", "WARNING", "ERROR")):
        self.logs = deque()  # deque seems to be better large logs
        self.console_levels = console_levels
        
    def log(self, message, level="DEBUG"):
        """Add a log message, timestamp and level"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.logs.append([timestamp, level, message])
        if level in self.console_levels:
            print(message)
        
    def get_logs(self):
        """Return all logs"""
        return list(self.logs)

class ExcelWriter:
    """Write data to Excel"""
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.workbook = None
        self.output_path = None
        
    def save_data(self, log_data, debug_logs=None):
        """Save data to an Excel file"""
        try:
            # Start a workbook and set up file parameters
            self.workbook = openpyxl.Workbook()
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            base_filename = self.config['config'].get('title', 'datalayer')
            output_folder = get_output_folder(self.config, self.logger)
            self.output_path = output_folder / f'{base_filename}_{timestamp}.xlsx'
            
            # Write each sequence to a separate sheet
            first_sequence = True
            for sequence_name, sequence_data in log_data.items():
                if first_sequence:
                    sheet = self.workbook.active
                    sheet.title = sequence_name[:31]
                    first_sequence = False
                else:
                    sheet = self.workbook.create_sheet(title=sequence_name[:31])
                
                self._write_sequence_data(sheet, sequence_data)
            
            # Add debug logs if enabled
            if debug_logs and self.config['config'].get('debug_mode', False):
                self.logger.log("Creating debug log sheet...")
                debug_sheet = self.workbook.create_sheet(title="debug_log")
                self._write_debug_logs(debug_sheet, debug_logs)
            
            # Save workbook
            self.workbook.save(self.output_path)
            # self.logger.log(f"Data successfully saved to Excel: {self.output_path}", "INFO")
            return str(self.output_path)
            
        except Exception as e:
            error_msg = f"Failed to save to Excel: {str(e)}"
            self.logger.log(error_msg, "ERROR")
            raise Exception(error_msg)
    
    def _write_sequence_data(self, sheet, data):
        """Write sequence data to a sheet"""
        
        # Headers
        headers = ["Step", "Event", "Timestamp", "URL", "Event Data", "Valid", "Error Details"]
        sheet.append(headers)
        
        # Set column & row sizes
        sheet.column_dimensions['A'].width = 20  # Step
        sheet.column_dimensions['B'].width = 20  # Event
        sheet.column_dimensions['C'].width = 20  # Timestamp
        sheet.column_dimensions['D'].width = 50  # URL
        sheet.column_dimensions['E'].width = 100  # Event Data
        sheet.column_dimensions['F'].width = 20  # Valid
        sheet.column_dimensions['G'].width = 100  # Error Details
        
        sheet.sheet_format.defaultRowHeight = 20 # Default row height for all rows
        sheet.row_dimensions[1].height = 15 # Header row height
        
        # Write data
        for entry in data:
            sheet.append(list(entry))
        
        # Apply text wrapping to Event Data column
        for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
    
    def _write_debug_logs(self, sheet, debug_logs):
        """If enabled, write debug logs to an additional sheet"""
        
        # Write headers
        headers = ["Timestamp", "Level", "Message"]
        sheet.append(headers)
        
        # Set column & row sizes
        sheet.column_dimensions['A'].width = 20  # Timestamp
        sheet.column_dimensions['B'].width = 10  # Level
        sheet.column_dimensions['C'].width = 150  # Message
        sheet.sheet_format.defaultRowHeight = 20 # Default row height for all rows
        
        # Add logs
        for log in debug_logs:
            formatted_row = []
            for item in log:
                cell_value = str(item)
                # If cell starts with =, +, -, @, tab, or has common Excel triggers
                if cell_value.startswith(('=', '+', '-', '@', '\t')) or ',,' in cell_value:
                    # Prefix with single quote to force text format
                    cell_value = f"'{cell_value}"
                formatted_row.append(cell_value)
            sheet.append(formatted_row)
        
        # Apply text wrapping to Message column
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
                cell.data_type = 's'

class GoogleSheetsAuth:
    """Google Sheets authentication using OAuth 2.0"""
    
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
              'https://www.googleapis.com/auth/drive.file']
    
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.credentials = None
        
        # Get Google Sheets config from the TOML file (config.google_sheets)
        self.gs_config = self.config['config'].get('google_sheets', {})

        # Get and validate credentials paths
        self.credentials_path = self._get_credentials_path()
        self.token_path = self._get_token_path()
        self._validate_paths()
        
    def _get_credentials_path(self):
        """Get credentials.json path from a path or environment"""
        if self.gs_config.get('credentials_location') == 'env':
            path = os.getenv('GOOGLE_SHEETS_CREDENTIALS_PATH')
            if not path:
                raise ValueError("GOOGLE_SHEETS_CREDENTIALS_PATH environment variable not set")
            self.logger.log("Using credentials path from environment variable")
            return path
        else:
            # Default to file-based configuration
            path = self.gs_config.get('credentials_path', 'credentials.json')
            # If relative path, make it relative to config file location
            if not os.path.isabs(path):
                config_dir = os.path.dirname(os.path.abspath(self.config['_config_file_path']))
                path = os.path.join(config_dir, path)
            self.logger.log(f"Using credentials path from config: {path}")
            return path
            
    def _get_token_path(self):
        """Get token.pickle path from a path or environment"""
        if self.gs_config.get('token_location') == 'env':
            path = os.getenv('GOOGLE_SHEETS_TOKEN_PATH')
            if not path:
                raise ValueError("GOOGLE_SHEETS_TOKEN_PATH environment variable not set")
            self.logger.log("Using token path from environment variable")
            return path
        else:
            # Default to file-based configuration - same directory as credentials
            creds_dir = os.path.dirname(self.credentials_path)
            path = os.path.join(creds_dir, 'token.pickle')
            self.logger.log(f"Using token path: {path}")
            return path
            
    def _validate_paths(self):
        """Validate that credential paths exist and are accessible"""
        # Check credentials.json
        if not os.path.exists(self.credentials_path):
            raise FileNotFoundError(
                f"Credentials file not found at: {self.credentials_path}\n"
                "If you don't have the credentials.json file, please follow the setup instructions to configure Google Sheets integration."
            )
            
        # Check token directory is writable if token doesn't exist yet
        token_dir = os.path.dirname(self.token_path)
        if not os.path.exists(self.token_path):
            if not os.path.exists(token_dir):
                try:
                    os.makedirs(token_dir)
                except Exception as e:
                    raise PermissionError(
                        f"Cannot create token directory at: {token_dir}\n"
                        f"Error: {str(e)}"
                    )
            elif not os.access(token_dir, os.W_OK):
                raise PermissionError(
                    f"Token directory is not writable: {token_dir}\n"
                    "Please ensure you have write permissions."
                )
        
    def authenticate(self):
        """OAuth authentication flow"""
        try:
            if os.path.exists(self.token_path):
                with open(self.token_path, 'rb') as token:
                    self.credentials = pickle.load(token)
                    
            # If there are no (valid) credentials available, let the user log in
            if not self.credentials or not self.credentials.valid:
                if self.credentials and self.credentials.expired and self.credentials.refresh_token:
                    self.logger.log("Refreshing Google Sheets access token...")
                    self.credentials.refresh(Request())
                else:
                    self.logger.log("Starting new Google Sheets authentication flow...", "INFO")
                    flow = InstalledAppFlow.from_client_secrets_file(
                        self.credentials_path,
                        self.SCOPES
                    )
                    self.credentials = flow.run_local_server(port=0)
                    
                # Save the credentials
                token_dir = os.path.dirname(self.token_path)
                if not os.path.exists(token_dir):
                    os.makedirs(token_dir)
                    
                with open(self.token_path, 'wb') as token:
                    pickle.dump(self.credentials, token)
            self.service = build('sheets', 'v4', credentials=self.credentials)
            return self.service        
            
        except Exception as e:
            self.logger.log(f"Authentication error: {str(e)}", "ERROR")
            raise        
    
class GoogleSheetsWriter:
    """Writing data to Google Sheets"""
    
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.spreadsheet_id = None
        self.service = None
        self.spreadsheet_url = None
        
    def save_data(self, log_data, debug_logs=None):
        """Main method to save data to Google Sheets"""
        try:
            # Initialize authentication
            auth = GoogleSheetsAuth(self.config, self.logger)
            auth.authenticate()
            self.service = auth.service
            
            # Create new spreadsheet
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            base_filename = self.config['config'].get('title', 'datalayer')
            sheet_title = f"{base_filename}_{timestamp}"
            
            self.logger.log(f"Creating Google Sheet: {sheet_title}")
            spreadsheet = self.service.spreadsheets().create(
                body={'properties': {'title': sheet_title}}
            ).execute()
            self.spreadsheet_id = spreadsheet['spreadsheetId']

            # Move to folder if a directory is specified
            if 'folder_id' in self.config['config'].get('google_sheets', {}):
                self.logger.log("Moving spreadsheet to a specified folder")
                self._move_to_folder()
            
            # Write each sequence to a separate sheet
            first_sequence = True
            for sequence_name, sequence_data in log_data.items():
                self._write_sequence_data(sequence_name, sequence_data)
                
                if first_sequence:
                    # Remove a default empty sheet after creating the first non-empty sheet
                    try:
                        spreadsheet_info = self.service.spreadsheets().get(
                            spreadsheetId=self.spreadsheet_id
                        ).execute()
                        sheet1_id = spreadsheet_info['sheets'][0]['properties']['sheetId']
                        
                        delete_request = {
                            'deleteSheet': {
                                'sheetId': sheet1_id
                            }
                        }
                        self.service.spreadsheets().batchUpdate(
                            spreadsheetId=self.spreadsheet_id,
                            body={'requests': [delete_request]}
                        ).execute()
                    except Exception as e:
                        self.logger.log(f"Warning: Could not remove default Sheet1: {str(e)}", "WARNING")
                    
                    first_sequence = False
            
            # Add debug logs if enabled
            if debug_logs and self.config['config'].get('debug_mode', False):
                self._write_debug_logs(debug_logs)
            
            self.spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{self.spreadsheet_id}"
            #self.logger.log(f"Data successfully saved to Google Sheets: {self.spreadsheet_url}", "INFO")
            return self.spreadsheet_url
        
        except HttpError as e:
            if e.resp.status == 429 or 'quotaExceeded' in str(e):
                error_msg = "Google Sheets API quota exceeded. Try again later or use Excel file as the output."
                self.logger.log(error_msg, "ERROR")
                raise Exception(error_msg)
            raise    
        
        except Exception as e:
            self.logger.log(f"Error saving to Google Sheets: {str(e)}", "ERROR")
            raise
            
    def _move_to_folder(self):
        """Move the spreadsheet to a specified Google Drive folder"""
        try:
            folder_id = self.config['config']['google_sheets']['folder_id']
            
            # Get credentials
            auth = GoogleSheetsAuth(self.config, self.logger)
            auth.authenticate()
            credentials = auth.credentials
            
            # Create Drive service with authenticated credentials
            drive_service = build('drive', 'v3', credentials=credentials)
            
            # Get the file's current parents
            file = drive_service.files().get(
                fileId=self.spreadsheet_id,
                fields='parents'
            ).execute()
            
            # Move file to the new folder
            previous_parents = ",".join(file.get('parents', []))
            drive_service.files().update(
                fileId=self.spreadsheet_id,
                addParents=folder_id,
                removeParents=previous_parents,
                fields='id, parents'
            ).execute()
            
        except HttpError as e:
            if e.resp.status == 429 or 'quotaExceeded' in str(e):
                error_msg = "Google Drive API quota exceeded. Try again later or use Excel file as the output."
                self.logger.log(error_msg, "ERROR")
                raise Exception(error_msg)
            raise    
                    
        except Exception as e:
            self.logger.log(f"Warning: Could not move file to the specified folder: {str(e)}", "WARNING")
    
    def _write_sequence_data(self, sequence_name, data):
        """Write sequence data to a sheet"""
        try:
            # Prepare headers and values
            headers = ["Step", "Event", "Timestamp", "URL", "Event Data", "Valid", "Error Details"]
            values = [headers] + data
            
            # Create new sheet
            sheet_id = self._create_sheet(sequence_name)
            
            # Update values
            body = {'values': values}
            self.service.spreadsheets().values().update(
                spreadsheetId=self.spreadsheet_id,
                range=f"{sequence_name}!A1",
                valueInputOption='RAW',
                body=body
            ).execute()
            
            # Apply formatting
            self._apply_formatting(sheet_id, len(values), len(headers))

        except HttpError as e:
            if e.resp.status == 429 or 'quotaExceeded' in str(e):
                error_msg = "Google Sheets API quota exceeded. Try again later or use Excel file as the output."
                self.logger.log(error_msg, "ERROR")
                raise Exception(error_msg)
            raise    

        except Exception as e:
            self.logger.log(f"Error writing sheet {sequence_name}: {str(e)}", "ERROR")
            raise
    
    def _create_sheet(self, sheet_name):
        """Create a new sheet and return its ID"""
        try:
            request = {
                'addSheet': {
                    'properties': {
                        'title': sheet_name[:31]  # Sheet names are limited to 31 chars
                    }
                }
            }
            response = self.service.spreadsheets().batchUpdate(
                spreadsheetId=self.spreadsheet_id,
                body={'requests': [request]}
            ).execute()
            return response['replies'][0]['addSheet']['properties']['sheetId']

        except HttpError as e:
            if e.resp.status == 429 or 'quotaExceeded' in str(e):
                error_msg = "Google Sheets API quota exceeded. Try again later or use Excel file as the output."
                self.logger.log(error_msg, "ERROR")
                raise Exception(error_msg)
            raise    

        except Exception as e:
            self.logger.log(f"Error creating sheet {sheet_name}: {str(e)}", "ERROR")
            raise
    
    def _apply_formatting(self, sheet_id, row_count, col_count):
        """Apply formatting to the sheet"""
        try:
            requests = [
                # Format header row
                {
                    'repeatCell': {
                        'range': {
                            'sheetId': sheet_id,
                            'startRowIndex': 0,
                            'endRowIndex': 1
                        },
                        'cell': {
                            'userEnteredFormat': {
                                'backgroundColor': {'red': 0.9, 'green': 0.9, 'blue': 0.9},
                                'textFormat': {'bold': True}
                            }
                        },
                        'fields': 'userEnteredFormat(backgroundColor,textFormat)'
                    }
                },
                # Adjust column widths
                {
                    'autoResizeDimensions': {
                        'dimensions': {
                            'sheetId': sheet_id,
                            'dimension': 'COLUMNS',
                            'startIndex': 0,
                            'endIndex': col_count
                        }
                    }
                },
                # Set row height
                {
                    'updateDimensionProperties': {
                        'range': {
                            'sheetId': sheet_id,
                            'dimension': 'ROWS',
                            'startIndex': 0,
                            'endIndex': row_count
                        },
                        'properties': {
                            'pixelSize': 30  # height
                        },
                        'fields': 'pixelSize'
                    }
                }
            ]
            
            self.service.spreadsheets().batchUpdate(
                spreadsheetId=self.spreadsheet_id,
                body={'requests': requests}
            ).execute()
            
        except Exception as e:
            self.logger.log(f"Warning: Could not apply formatting: {str(e)}", "WARNING")
    
    def _write_debug_logs(self, debug_logs):
        """Write debug logs to a separate sheet"""
        try:
            # Create debug log sheet
            sheet_name = "debug_log"
            sheet_id = self._create_sheet(sheet_name)
            
            # Prepare headers and values
            headers = ["Timestamp", "Level", "Message"]
            values = [headers] + debug_logs
            
            # Update values
            body = {'values': values}
            self.service.spreadsheets().values().update(
                spreadsheetId=self.spreadsheet_id,
                range=f"{sheet_name}!A1",
                valueInputOption='RAW',
                body=body
            ).execute()
            
            # Apply formatting
            self._apply_formatting(sheet_id, len(values), len(headers))

        except HttpError as e:
            if e.resp.status == 429 or 'quotaExceeded' in str(e):
                error_msg = "Google Sheets API quota exceeded. Try again later or use Excel file as the output."
                self.logger.log(error_msg, "ERROR")
                raise Exception(error_msg)
            raise

        except Exception as e:
            self.logger.log(f"Warning: Could not write debug logs: {str(e)}", "WARNING")

##### FUNCTIONS

def clean_error_message(error):
    """Clean error message by removing stacktrace and technical details"""
    error_str = str(error)
    # Get first line of error message or everything before stacktrace
    cleaned = error_str.split('\n')[0].split('Stacktrace')[0].strip()
    if not cleaned:  # If empty after cleaning, use a generic message
        cleaned = f"Browser error occurred: {error.__class__.__name__}"
    return cleaned

def parse_validation_from_toml(config, logger):
    """
    Parse validation rules from TOML config.
    Returns a dictionary of event_name -> validation_rules.
    """
    validation_rules = {}
    
    if 'validation' not in config:
        return validation_rules
        
    for event_name, event_config in config['validation'].items():
        if 'code' not in event_config:
            continue
            
        code_str = event_config['code']
        # Remove whitespace and newlines
        code_str = code_str.strip()
        if not (code_str.startswith('{') and code_str.endswith('}')):
            code_str = "{" + code_str + "}"
        # Remove comments and clean up
        code_lines = [line.split('#')[0].strip() for line in code_str.split('\n')]
        code_str = ' '.join(code_lines)
        
        try:
            rule_dict = parse_validation_code_block(code_str, logger)
            validation_rules[event_name] = rule_dict
        except ValueError as e:
            print(f"Error parsing validation rule for {event_name}: {str(e)}")
    
    return validation_rules

def parse_validation_code_block(code_str, logger):
    """
    A parser for validation blocks from TOML files.
    Returns a dictionary structure.
    """
    def tokenize(s):
        """
        Tokenize string into: { } [ ] : , /regex/ <type> strings
        or unquoted keys. Returns a list of tokens.
        """
        tokens = []
        i = 0
        while i < len(s):
            c = s[i]

            # Skip whitespace
            if c.isspace():
                i += 1
                continue

            # Single character tokens
            if c in '{}[]:,':
                tokens.append(c)
                i += 1
                continue

            # Regex patterns
            if s.startswith('/', i):
                j = s.find('/', i+1)
                if j == -1:
                    raise ValueError("Unclosed '/' in regex pattern")
                regex_token = s[i:j+1]
                try:
                    re.compile(regex_token[1:-1])  # Validate it's a valid regex
                    tokens.append(regex_token)
                    i = j+1
                    continue
                except re.error:
                    raise ValueError(f"Invalid regex pattern: {regex_token}")

            # Type specifiers
            if s.startswith('<', i):
                j = s.find('>', i+1)
                if j == -1:
                    raise ValueError("Unclosed '< >' for type")
                type_token = s[i:j+1]
                # Validate allowed type specifiers
                if type_token.lower() not in allowed_validation_types:
                    raise ValueError(f"Invalid type specifier '{type_token}'. Must be one of: {', '.join(allowed_validation_types)}")
                tokens.append(type_token)
                i = j+1
                continue

            # Quoted strings
            if s.startswith('"', i) or s.startswith("'", i):
                quote_char = s[i]
                j = i + 1
                buff = []
                escaped = False
                
                while j < len(s):
                    curr_char = s[j]
                    
                    if escaped:
                        # Handle escaped character
                        if curr_char == quote_char or curr_char == '\\':
                            buff.append(curr_char)
                        else:
                            buff.append('\\' + curr_char)
                        escaped = False
                        j += 1
                        continue
                        
                    if curr_char == '\\':
                        escaped = True
                        j += 1
                        continue
                        
                    if curr_char == quote_char:
                        break
                        
                    buff.append(curr_char)
                    j += 1
                    
                if j >= len(s) or s[j] != quote_char:
                    raise ValueError(f"Unterminated string starting at position {i}")
                    
                tokens.append(''.join(buff))
                i = j + 1
                continue

            # Unquoted keys/values 
            special = set('{}[]:,"\'/ \t\n\r')
            buff = []
            while i < len(s) and s[i] not in special:
                buff.append(s[i])
                i += 1
            tokens.append(''.join(buff))

        return tokens

    # Functions to parse objects, lists, etc. using tokens
    tokens = tokenize(code_str)
    index = 0  # global pointer to token list

    def peek():
        return tokens[index] if index < len(tokens) else None

    def consume(expected=None):
        nonlocal index
        if index >= len(tokens):
            return None
        t = tokens[index]
        index += 1
        if expected and t != expected:
            raise ValueError(f"Expected '{expected}' but got '{t}'")
        return t

    def parse_value():
        t = peek()
        if t == '{':
            return parse_object(logger)
        elif t == '[':
            return parse_array()
        elif t and t.startswith('/'):  # /regex/
            consume()
            return t  # return as is
        elif t and t.startswith('<'):  # <type>
            consume()
            return t  # return as is
        else:
            # treat anything else as string
            return consume()

    def parse_object(logger):
        obj = {}
        consume('{')
        first = True
        while True:
            t = peek()
            if t == '}':
                consume('}')
                break
            if not first and t == ',':
                consume(',')
                if peek() == '}':
                    consume('}')
                    break
            # key
            key = parse_value() 
            if not isinstance(key, str):
                logger.log(f"⚠️  Validation key should be a string: {str(key)}.\nPlease review your configuration file.", "ERROR")
            # Handle required field marker (!)
            required = False
            if key.startswith('!'):
                required = True
                key = key[1:]  # Remove the ! prefix
            
            # Handle quoted keys
            if key.startswith('"') and key.endswith('"') or key.startswith("'") and key.endswith("'"):
                key = key[1:-1]  # Remove quotes
                
            # colon
            if peek() == ':':
                consume(':')
            # value
            val = parse_value()
            
            # Store with metadata for required fields
            if required:
                obj[f"!{key}"] = val
            else:
                obj[key] = val
            first = False
        return obj

    def parse_array():
        """
        Parse a list structure from the tokenized input.
        Returns a Python list containing the parsed elements.
        """
        arr = []
        consume('[')
        first = True
        while True:
            t = peek()
            if t == ']':
                consume(']')
                break
            if not first and t == ',':
                consume(',')
                if peek() == ']':
                    consume(']')
                    break
            val = parse_value()
            arr.append(val)
            first = False
        return arr

    # Start main parsing
    if not tokens:
        return {}
    result = parse_object(logger) if tokens[0] == '{' else parse_object(logger)
    return result

def validate_sequence(config, logger):
    """
    Validate the entire sequence configuration with detailed checks.
    Raises ValueError with specific messages if validation fails.
    """
    # Check for required top-level sections
    required_sections = ['config', 'step', 'sequence']
    missing_sections = [section for section in required_sections if section not in config]
    if missing_sections:
        raise ValueError(f"Missing required configuration sections: {', '.join(missing_sections)}")
    
    # Validate user agents
    if not config['config'].get('user_agents'):
        logger.log(f"No user agent specified in the config file, using the default user agent")
        config['config']['user_agents'] = [ 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36' ]
    if not isinstance(config['config']['user_agents'], list) or not config['config']['user_agents']:
        raise ValueError("config.user_agents must be a non-empty list")
    if not all(isinstance(ua, str) for ua in config['config']['user_agents']):
        raise ValueError("All user agents must be strings - use quotation marks")
    
    # Validate config section
    # required_config = [] # if any config field is obligatory, add it here
    # missing_config = [field for field in required_config if field not in config['config']]
    # if missing_config:
    #     raise ValueError(f"Missing required config fields: {', '.join(missing_config)}")
    
    # Validate output configuration
    output_destination = config['config'].get('output_destination', 'excel')
    if output_destination not in ['excel', 'google_sheets']:
        logger.log(f"Warning: Invalid output_destination '{output_destination}' - must be 'excel' or 'google_sheets'. Using 'excel' as default.", "WARNING")
        config['config']['output_destination'] = 'excel'
        
    # Validate Google Sheets configuration when it's selected as output
    if output_destination == 'google_sheets':
        if 'google_sheets' not in config['config']:
            raise ValueError("google_sheets configuration required when output_destination is 'google_sheets'")
            
        gs_config = config['config']['google_sheets']
        
        # Validate credentials configuration
        credentials_location = gs_config.get('credentials_location', 'file')
        if credentials_location not in ['file', 'env']:
            logger.log(f"Warning: Invalid credentials_location '{credentials_location}' - must be 'file' or 'env'. Using 'file' as default.", "WARNING")
            gs_config['credentials_location'] = 'file'
            
        if credentials_location == 'file' and not gs_config.get('credentials_path'):
            raise ValueError("credentials_path required when credentials_location is 'file' - should be a string in quotation marks, e.g. 'credentials.json'")
            
        if credentials_location == 'env' and not os.getenv('GOOGLE_SHEETS_CREDENTIALS_PATH'):
            raise ValueError("GOOGLE_SHEETS_CREDENTIALS_PATH environment variable required when credentials_location is 'env'")
            
        # Validate token configuration
        token_location = gs_config.get('token_location', 'file')
        if token_location not in ['file', 'env']:
            logger.log(f"Warning: Invalid token_location '{token_location}' - must be 'file' or 'env'. Using 'file' as default.", "WARNING")
            gs_config['token_location'] = 'file'
            
        # Folder ID is optional but must be string if present
        folder_id = gs_config.get('folder_id')
        if folder_id is not None and not isinstance(folder_id, str):
            raise ValueError("folder_id must be a string in quotation marks, e.g. '1234567890abcdef'") 

    if 'output_folder' in config['config']:
        output_folder = config['config']['output_folder']
        if not isinstance(output_folder, str):
            raise ValueError("output_folder must be a string in quotation marks, e.g. 'results' or '/path/to/folder'")

    # Get set of defined steps
    steps_defined = set(config['step'].keys())
    if not steps_defined:
        raise ValueError("No steps defined in configuration")
    
    # Validate each step definition
    for step_name, step in config['step'].items():
        if 'type' not in step:
            raise ValueError(f"Step '{step_name}' missing required 'type' field")
            
        # Validate step type-specific requirements
        if step['type'] == 'visit':
            # URL is optional for visit steps
            if 'url' in step:
                # If URL is present, validate its format
                if isinstance(step['url'], list):
                    if not all(isinstance(u, str) for u in step['url']):
                        raise ValueError(f"All URLs in step '{step_name}' must be strings")
                elif not isinstance(step['url'], str):
                    raise ValueError(f"URL in step '{step_name}' must be a string or list of strings")
                    
        elif step['type'] == 'click':
            if 'clicks' not in step:
                raise ValueError(f"Click step '{step_name}' missing required 'clicks' list")
            if not isinstance(step['clicks'], list):
                raise ValueError(f"Clicks in step '{step_name}' must be a list")
            for i, click in enumerate(step['clicks']):
                if not isinstance(click, dict):
                    raise ValueError(f"Click {i} in step '{step_name}' must be a dictionary (in curly brackets), e.g. " + "{ selector = 'a.button_purchase' }")
                if not ('xpath' in click or 'selector' in click):
                    raise ValueError(f"Click {i} in step '{step_name}' missing either 'xpath' or 'selector'")
                # Validate delay_after if present
                if 'delay_after' in click and not isinstance(click['delay_after'], (int, float)):
                    raise ValueError(f"delay_after in click {i} of step '{step_name}' must be a number - without quotation marks, e.g. delay_after = 2")
                    
        elif step['type'] == 'form':
            if 'fields' not in step:
                raise ValueError(f"Form step '{step_name}' missing required 'fields' list")
            if not isinstance(step['fields'], list):
                raise ValueError(f"Fields in step '{step_name}' must be a list, in square brackets []")
            if 'submit_button' not in step:
                raise ValueError(f"Form step '{step_name}' missing required 'submit_button'. See the documentation for more details'")
            for i, field in enumerate(step['fields']):
                if not isinstance(field, dict):
                    raise ValueError(f"Field {i} in form step '{step_name}' must be a dictionary (in curly brackets), e.g. " + "{ selector = '#FirstNameInput', value = 'John' }")
                if not ('xpath' in field or 'selector' in field):
                    raise ValueError(f"Field {i} in form step '{step_name}' missing either 'xpath' or 'selector'")  
        
        elif step['type'] == 'scroll':
            # Must have exactly one of: xpath, selector, pixels, or percentage
            scroll_params = ['xpath', 'selector', 'pixels', 'percentage']
            present_params = [param for param in scroll_params if param in step]
            
            if not present_params:
                raise ValueError(f"Scroll step '{step_name}' must specify one of: xpath, selector, pixels, or percentage")
                
            if len(present_params) > 1:
                raise ValueError(f"Scroll step '{step_name}' can only specify one of: xpath, selector, pixels, or percentage")
                
            # If using pixels or percentage, validate they're numbers
            if 'pixels' in step:
                if not isinstance(step['pixels'], (int, float)):
                    raise ValueError(f"Pixels in scroll step '{step_name}' must be a number, without quotation marks, e.g. pixels = 100")
                if step['pixels'] <= 0:
                    raise ValueError(f"Pixels in scroll step '{step_name}' must be positive (more than 0)")
                    
            if 'percentage' in step:
                if not isinstance(step['percentage'], (int, float)):
                    raise ValueError(f"Percentage in scroll step '{step_name}' must be a number between 0 and 100, without quotation marks and wothout % sign, e.g. percentage = 75")
                if not 0 <= step['percentage'] <= 100:
                    raise ValueError(f"Percentage in scroll step '{step_name}' must be between 0 and 100")
        
        else:
            raise ValueError(f"Unknown step type '{step['type']}' in step '{step_name}'")
            
        # Validate step parameters
        if 'delay_after' in step:
            if not isinstance(step['delay_after'], (int, float)):
                raise ValueError(f"delay_after in step '{step_name}' must be a number, without quotation marks, e.g. delay_after = 2")
            if step['delay_after'] < 0:
                raise ValueError(f"delay_after in step '{step_name}' cannot be negative")
    
    # Validate sequences
    if not config['sequence']:
        raise ValueError("No sequences defined in configuration")
        
    for sequence_name, sequence in config['sequence'].items():
        if 'steps' not in sequence:
            raise ValueError(f"Sequence '{sequence_name}' missing required 'steps' list")
            
        if not isinstance(sequence['steps'], list):
            raise ValueError(f"Steps in sequence '{sequence_name}' must be a list, in square brackets []")
            
        if not sequence['steps']:
            raise ValueError(f"Sequence '{sequence_name}' contains no steps")
            
        # Check for undefined steps in sequence
        unknown_steps = [step for step in sequence['steps'] if step not in steps_defined]
        if unknown_steps:
            raise ValueError(f"Sequence '{sequence_name}' contains undefined steps: {', '.join(unknown_steps)}")
    
    # Validate delays and timeouts
    if 'default_timeout' in config['config']:
        if not isinstance(config['config']['default_timeout'], (int, float)):
            raise ValueError("default_timeout must be a number, e.g. default_timeout = 10")
        if config['config']['default_timeout'] <= 0:
            raise ValueError("default_timeout must be positive (higher than zero)")
            
    if 'default_delay' in config['config']:
        if not isinstance(config['config']['default_delay'], (int, float)):
            raise ValueError("default_delay must be a number, e.g. default_delay = 2")
        if config['config']['default_delay'] < 0:
            raise ValueError("default_delay cannot be negative, zero or higher")
    
    # All validations passed
    return True

def load_config(config_path, logger):
    """Load configuration from a TOML file"""
    try:
        logger.log(f"Loading configuration from {config_path}")
        with open(config_path, 'r') as file:
            config = toml.load(file)
        
        config['validation'] = parse_validation_from_toml(config, logger)
             
        # track_events configuration (all if not specified)
        if 'config' in config:
            track_events = config['config'].get('track_events')
            if track_events is None or (isinstance(track_events, list) and not track_events):
                logger.log("No track_events specified - will track all events", "INFO")
                config['config']['track_events'] = None

        validate_sequence(config, logger)
        logger.log("Configuration validation passed")
        return config
    
    except Exception as e:
        logger.log(f"Error loading configuration: {clean_error_message(e)}", "ERROR")
        sys.exit(1)

def initialize_browser(config, logger):
    """Initialize browser"""
    try:
        logger.log("Initializing browser with user agent settings")
        browser_options = webdriver.ChromeOptions()
        user_agent = random.choice(config['config']['user_agents'])
        
        if config['config'].get('include_selenium_info', False):
            user_agent += " Selenium"
            
        browser_options.add_argument(f'user-agent={user_agent}')
        browser_options.add_argument("--log-level=3") # Disable logging for webdriver
        browser_options.add_experimental_option('excludeSwitches', ['enable-logging']) # Disable DevTools logs
        browser_options.add_argument("--disable-usb") # fixes some error logs; remove if you really need USB
        
        # Add request blocking if configured
        block_rules = []
        
        # Built-in blockers for common services
        if config['config'].get('block_ga4', False):
            block_rules.extend([
                "*google-analytics.com/*",
                "*analytics.google.com/*",
                "*googletagmanager.com/gtag/*"
            ])
            
        if config['config'].get('block_gtm', False):
            block_rules.append("*googletagmanager.com/*")
            
        if config['config'].get('block_piwik', False):
            block_rules.append("*piwik.pro/*")
            
        # Add custom domain blocking
        if 'block_domains' in config['config']:
            custom_domains = config['config']['block_domains']
            if isinstance(custom_domains, list):
                for domain in custom_domains:
                    if isinstance(domain, str):
                        # Add both with and without www prefix
                        clean_domain = domain.replace('http://', '').replace('https://', '').rstrip('/')
                        block_rules.extend([
                            f"*://{clean_domain}/*",
                            f"*://*.{clean_domain}/*"
                        ])
                        logger.log(f"Added custom domain blocking for: {clean_domain}")
                    else:
                        logger.log(f"Warning: Invalid domain format in block_domains: {domain}", "ERROR")
            else:
                logger.log("Warning: block_domains must be a list", "ERROR")
            
        if block_rules:
            browser_options.add_argument('--enable-features=NetworkService')
            browser_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
            browser_options.add_experimental_option('excludeSwitches', ['enable-automation'])
            
            prefs = {
                "profile.default_content_settings.popups": 0,
                "profile.block_third_party_cookies": False
            }
            browser_options.add_experimental_option("prefs", prefs)
        
        browser = webdriver.Chrome(options=browser_options)
        if block_rules:
            browser.execute_cdp_cmd('Network.enable', {})
            browser.execute_cdp_cmd('Network.setBlockedURLs', {"urls": block_rules})
        
        logger.log("Browser initialized successfully")
        return browser
        
    except Exception as e:
        logger.log(f"Failed to initialize browser: {clean_error_message(e)}", "ERROR")
        sys.exit(1)

def get_element_locator(params, config):
    """
    Determine whether to use XPath or CSS selector for finding elements, then choose the right Selenium's By strategy.
    Uses XPath if both are present.
    """
    if 'xpath' in params:
        return (By.XPATH, params['xpath'])
    elif 'selector' in params:
        return (By.CSS_SELECTOR, params['selector'])
    else:
        raise ValueError("No valid selector found in parameters - must be either 'xpath' or 'selector'")

def has_dimensions(element):
    """Validate if element has non-zero dimensions"""
    try:
        return element.size['height'] > 0 and element.size['width'] > 0
    except:
        return False

def is_element_clickable(element):
    """Detailed check for an element to click"""
    try:
        # Check if element or its parents have 'hidden' attribute
        current = element
        while current:
            if current.get_attribute('hidden') is not None:
                return False
            if current.tag_name == 'body':
                break
            current = current.find_element(By.XPATH, '..')
        
        return element.is_enabled()
    except:
        return False

def wait_for_element(browser, params, config, logger):
    """Wait for element (with default randomized selection when multiple elements match)"""
    timeout = config['config'].get('default_timeout', 10) # default timeout
    max_elements = 50  # Threshold for "too many elements" warning
    
    try:
        by_strategy, selector = get_element_locator(params, config) # get_element_locator returns a tuple with two values
        logger.log(f"Waiting for elements matching: {selector}")
        
        # First wait for presence
        WebDriverWait(browser, timeout).until(
            EC.presence_of_element_located((by_strategy, selector))
        )
        
        # Get all matching elements
        elements = browser.find_elements(by_strategy, selector)
        if not elements:
            raise Exception(f"No elements found matching: {selector}")
            
        # Quick filter exclude elements with 0 width/height
        candidates = [e for e in elements if has_dimensions(e)]
        total_matches = len(candidates)
        logger.log(f"{total_matches} out of {len(elements)} matches qualified")
        
        if total_matches > max_elements:
            warning_msg = f"Warning: Selector '{selector}' matches {total_matches} elements - consider using a more specific selector"
            logger.log(warning_msg, "ERROR")
        
        if not candidates:
            raise Exception(f"No visible elements found matching: {selector}")
        
        # Try up to 5 random elements
        for attempt in range(5):
            if len(candidates) == 0:
                raise Exception("No more candidates available after failed attempts")
                
            # Pick a random element
            element = random.choice(candidates)
            
            # Detailed check for this element only
            if not is_element_clickable(element):
                candidates.remove(element)
                logger.log(f"🔎 Selected element not clickable, trying another ({len(candidates)} remaining)", "INFO")
                continue
            
            # If element needs scrolling
            if not element.is_displayed():
                logger.log("🔎 Selected element not in viewport, scrolling into view", "INFO")
                browser.execute_script("""
                    arguments[0].scrollIntoView({
                        block: 'center',
                        behavior: 'instant'
                    });
                """, element)
                time.sleep(0.5)
            
            # Final check for clickability using the specific element
            try:
                WebDriverWait(browser, 3).until(  # Short timeout for final check
                    lambda driver: element.is_displayed() and element.is_enabled()
                )
                logger.log("🎯 Element is now visible and clickable", "INFO")
                return element
            except:
                candidates.remove(element)
                logger.log("🔎 Element is not clickable after scroll, trying another one", "INFO")
                continue
                
        if total_matches > max_elements:
            raise Exception(f"Selector '{selector}' matches too many elements ({total_matches}). Could not find clickable element after 5 attempts. Consider using a more specific selector")
        else:
            raise Exception("Could not find clickable element after 5 attempts")
        
    except Exception as e:
        page_source = browser.page_source # Get page source for debugging
        if selector in page_source:
            logger.log(f"Element found in page source but not interactable", "ERROR")
        else:
            logger.log(f"Element not found in the page source", "ERROR")
        raise Exception(f"Error: Element not found or not clickable: {selector}")
    
def inject_css(browser, config, logger):
    """Inject CSS rules to hide specified elements"""
    css_elements_to_hide = config['config'].get('css_elements_to_hide', [])
    css_rules = "\n".join([f"{selector} {{ display: none !important; }}" for selector in css_elements_to_hide])

    if css_rules:
        try:
            browser.execute_script(f"""
                if (!document.getElementById('custom-css-hide-elements')) {{
                    let styleSheet = document.createElement("style");
                    styleSheet.type = "text/css";
                    styleSheet.id = "custom-css-hide-elements";
                    styleSheet.innerText = `{css_rules}`;
                    document.head.appendChild(styleSheet);
                }}
            """)
            logger.log("CSS rules injected")
        except Exception as e:
            logger.log(f"Warning: Failed to inject CSS rules: {clean_error_message(e)}", "ERROR")

def sanitize_event_data(event_data):
    """Clean data from unnecessary elements"""
    try:
        if isinstance(event_data, dict):
            return {
                k: sanitize_event_data(v) 
                for k, v in event_data.items() 
                if not str(type(v)).find('selenium') > -1  # Skip Selenium objects
                and k != 'error'  # Skip error objects
                and k != 'trace'  # Skip stack traces
            }
        elif isinstance(event_data, list):
            return [sanitize_event_data(item) for item in event_data]
        elif isinstance(event_data, (str, int, float, bool, type(None))):
            return event_data
        else:
            return str(event_data)
    except Exception as e:
        return f"Error sanitizing data: {clean_error_message(e)}"

def start_monitoring_thread(browser, monitored_events, event_queue, stop_event, logger, config):
    """Monitor dataLayer thread"""
    processed_events = set()
    error_cooldown = 0
    last_valid_url = None
    
    # Get validation rules from config
    validation_rules = config.get('validation', {})

    # Get the initial URL from the first visit step in the first sequence
    initial_url = None
    try:
        first_sequence = next(iter(config['sequence'].values()))
        first_step_name = first_sequence['steps'][0]
        first_step = config['step'][first_step_name]
        if first_step['type'] == 'visit' and 'url' in first_step:
            url = first_step['url']
            initial_url = url[0] if isinstance(url, list) else url
    except Exception as e:
        logger.log(f"Warning: Could not get initial URL from config: {clean_error_message(e)}", "ERROR")
        initial_url = "Initializing page"
    
    while not stop_event.is_set():
        try:
            # Only proceed if browser is still responsive
            if not browser or error_cooldown > 0:
                error_cooldown = max(0, error_cooldown - 1)
                time.sleep(0.5)
                continue
                
            current_url = browser.current_url
            if not current_url.startswith('data:'):
                last_valid_url = current_url
                
            # Use last_valid_url if available, otherwise use initial_url from config
            url_to_log = last_valid_url or initial_url
                
            datalayer = browser.execute_script("return window.dataLayer || []")
            
            if not isinstance(datalayer, list):
                logger.log("Warning: dataLayer is not a list", "ERROR")
                continue
                
            for event in datalayer:
                if not isinstance(event, dict) or 'event' not in event:
                    continue
                    
                try:
                    sanitized_event = sanitize_event_data(event)
                    event_id = f"{event['event']}_{hash(json.dumps(sanitized_event, sort_keys=True))}"
                    
                    if event_id in processed_events or (monitored_events and event['event'] not in monitored_events):
                        continue
                    
                    valid_flag = None
                    error_details = None
                    is_valid = None
                    errors = []

                    if validation_rules:
                        event_name = event['event']
                        rule_for_event = validation_rules.get(event_name, {})
                        if rule_for_event:
                            is_valid, errors = validate_event(sanitized_event, rule_for_event)
                            valid_flag = "✔️" if is_valid else "❌"
                            error_details = errors if not is_valid else None
                        else:
                            valid_flag = "-"
                            error_details = None
                    else:
                        valid_flag = "-"
                        error_details = None

                    # Add validation result to the event record
                    event_queue.put({
                        'event_name': event['event'],
                        'event_data': sanitized_event,
                        'timestamp': datetime.now(),
                        'url': url_to_log,
                        'valid': valid_flag,
                        'error_details': error_details
                    })
                    
                    processed_events.add(event_id)

                    if is_valid is True:
                        logger.log(f"✅ Valid event: {event['event']}")
                    elif is_valid is False:
                        logger.log(f"⚠️  Invalid event: {event['event']}", "ERROR")
                    else:
                        logger.log(f"🟦 Not validated (no rule): {event['event']}")

                except Exception as inner_e:
                    logger.log(f"Error processing event: {clean_error_message(inner_e)}", "ERROR")
                    
        except Exception as e:
            error_msg = clean_error_message(e)
            logger.log(f"Error in monitoring thread: {error_msg}", "ERROR")
            error_cooldown = 10  # Add cooldown period after error
            
        time.sleep(0.1)

def process_queued_events(event_queue, log_data, current_step, logger, until_time=None):
    """Process events from queue until specified time"""
    while not event_queue.empty():
        try:
            event = event_queue.get_nowait()
            # If until_time is specified, only process events that occurred before it
            if until_time and event['timestamp'] > until_time:
                # Put the event back in the queue for the next step
                event_queue.put(event)
                break
                
            log_data.append([
                current_step,
                event['event_name'],
                event['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                event['url'],
                json.dumps(event['event_data'], indent=2) , # Added indentation for better formatting
                event['valid'],
                json.dumps(event.get('error_details', '-'), indent=2) if event.get('error_details') else "-"
            ])
        except Exception as e:
            logger.log(f"Error processing event from queue: {clean_error_message(e)}", "ERROR")

def perform_action(browser, action_type, params, config, logger):
    """Perform a single browser action - visit, click, form, scroll"""
    
    def do_click(element, browser, method="selenium"):
        """Provide different methods of clicking an element."""
        if method == "selenium":
            try:
                element.click()
            except Exception:
                browser.execute_script("arguments[0].click();", element)
        elif method == "js":
            browser.execute_script("arguments[0].click();", element)
        elif method == "action":
            ActionChains(browser).move_to_element(element).click().perform()
        else:
            raise ValueError(f"Unsupported click_method: {method}")
    
    try:
        if action_type != 'visit':
            logger.log(f"➡️  Current URL: {browser.current_url}", "INFO")

        if action_type == 'scroll':
            if 'selector' in params or 'xpath' in params:
                element = wait_for_element(browser, params, config, logger)
                logger.log(f"➡️ Scrolling to element", "INFO")
                browser.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)
            elif 'pixels' in params:
                scroll_amount = params['pixels']
                logger.log(f"➡️ Scrolling by {scroll_amount} pixels", "INFO")
                browser.execute_script(f"window.scrollBy(0, {scroll_amount});")
            elif 'percentage' in params:
                scroll_percentage = params['percentage']
                logger.log(f"➡️ Scrolling to {scroll_percentage}% of page", "INFO")
                browser.execute_script(f"""
                    let pageHeight = Math.max(
                        document.body.scrollHeight,
                        document.documentElement.scrollHeight
                    );
                    window.scrollTo(0, pageHeight * {scroll_percentage / 100});
                """)
            else:
                raise ValueError("Scroll step must specify either 'selector', 'xpath', 'pixels', or 'percentage'")
                
            # Short wait for scroll completion
            time.sleep(1)  # This is just for scroll animation, not the configured delay
            return f"Scrolled page successfully"

        elif action_type == 'visit':
            if 'url' not in params:
                logger.log("Step marked as page view without navigation")
                return "Page view step (no navigation)"
            else:
                url = random.choice(params['url']) if isinstance(params['url'], list) else params['url']
                final_url = url + "?bot=true" if config['config'].get('bot_info', False) else url
                browser.get(final_url)
                
                try:
                    WebDriverWait(browser, config['config'].get('default_timeout', 10)).until(
                        lambda driver: driver.execute_script('return document.readyState') == 'complete'
                    )
                    logger.log(f"➡️  Current URL: {browser.current_url}", "INFO")
                    logger.log("Page load completed")
                    inject_css(browser, config, logger)
                except Exception as e:
                    logger.log(f"Warning: Page load wait timed out: {clean_error_message(e)}", "ERROR")
                
                # Verify we're not on a blank/transitional page
                if browser.current_url.startswith('data:'):
                    logger.log("Warning: URL is not saved correctly for page navigation", "ERROR")
                
                return f"Visited URL: {final_url}"
            
        if action_type == 'click':
            clicks = params.get('clicks', [])
            if not clicks:
                raise ValueError("Click step must contain a 'clicks' list")
                    
            success_count = 0
            for i, click_params in enumerate(clicks):
                try:
                    # Only call wait_for_element once per click attempt
                    element = wait_for_element(browser, click_params, config, logger)
                    
                    try:
                        element.click()
                    except Exception as e:
                        browser.execute_script("arguments[0].click();", element)
                        
                    by_strategy, selector = get_element_locator(click_params, config)
                    logger.log(f"➡️  Clicked element {i+1}: {selector}" , "INFO")
                    success_count += 1
                    
                    # Handle delay between individual clicks
                    if i < len(clicks) - 1:  # Don't delay after last click
                        delay = click_params.get('delay_after', config['config'].get('default_delay', 1))
                        if delay > 0:
                            logger.log(f"Waiting {delay} seconds between clicks...")
                            time.sleep(delay)
                except Exception as click_error:
                    logger.log(f"Failed to click element {i+1}: {clean_error_message(click_error)}", "ERROR")
                    continue

            if success_count == 0:
                raise Exception("All clicks in step failed")
            elif success_count < len(clicks):
                return f"Completed {success_count} out of {len(clicks)} clicks"
            else:
                return "All clicks completed successfully"
            
        elif action_type == 'form':
            submit_method = params.get('submit_method', 'selenium')
            for field in params['fields']:
                element = wait_for_element(browser, field, config, logger)
                element.clear()  # clear input before filling in
                element.send_keys(field['value'])
            submit_button = wait_for_element(browser, {'xpath': params['submit_button']}, config, logger)
            do_click(submit_button, browser, submit_method) # click the submit button with the specified method
            return "Form submitted successfully"

            
    except Exception as e:
        error_msg = clean_error_message(e)
        raise Exception(error_msg)

def perform_sequence(browser, config, event_queue, sequence, logger):
    """Execute step sequence"""
    steps_definitions = config['step']
    default_delay = config['config'].get('default_delay', 1) # default delay is 1 second
    log_data = []

    logger.log(f"\n=== Starting sequence execution ===")

    # Exectuing step by step
    for i, step_name in enumerate(sequence['steps']):
        step = steps_definitions[step_name]
        is_final_step = i == len(sequence['steps']) - 1
        
        logger.log(f"\n=== Starting step: {step_name} ===", "INFO")
        
        try:
            # Inject CSS before any action (in perform_action there is an additional injection for visit steps after page load)
            inject_css(browser, config, logger)
            
            # Execute the step action
            result = perform_action(browser, step['type'], step, config, logger)
            logger.log(result)

            # Get step-level delay
            delay = step.get('delay_after', default_delay)
            
            # Handle delays based on step type and position
            if is_final_step:
                logger.log(f"Final step - waiting {delay} seconds for events...", "INFO")
            else:
                logger.log(f"Waiting {delay} seconds after {step['type']} step...")
                
            if delay > 0:
                time.sleep(delay)
                logger.log(f"Delay completed at {datetime.now().strftime('%H:%M:%S')}")
                
            # Calculate the cutoff time for events in this step
            step_end_time = datetime.now()
            
            # For the final step, don't use cutoff time
            if is_final_step:
                process_queued_events(event_queue, log_data, step_name, logger)
            else:
                process_queued_events(event_queue, log_data, step_name, logger, step_end_time)

            logger.log(f"🎉 Step {step_name} completed successfully", "INFO")
                
        except Exception as e:
            error_msg = clean_error_message(e)
            logger.log(f"Error in step {step_name}: {error_msg}", "ERROR")
            log_data.append([
                step_name,
                'Error',
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                browser.current_url,
                error_msg,
                "-",
                "-"
            ])

    return log_data

# For better efficiency, cache already compiled regex patterns for validation
compiled_regex_cache = {}

def get_compiled_pattern(pattern_str):
    """
    Returns a compiled regex pattern.
    If the pattern is not already compiled, it compiles it and stores it in the cache.
    """
    if pattern_str not in compiled_regex_cache:
        compiled_regex_cache[pattern_str] = re.compile(pattern_str)
    return compiled_regex_cache[pattern_str]
    
def validate_event(event: Dict[str, Any], rules: Dict[str, Any]) -> Tuple[bool, List[str]]:
    """
    Validates an event based on the provided rules.
    """
    errors = []
    
    type_checks = {
        "int": lambda v: isinstance(v, int) and not isinstance(v, bool),
        "float": lambda v: isinstance(v, (int, float)),
        "str": lambda v: isinstance(v, str),
    }

    
    def check_structure(data: Dict[str, Any], rule: Dict[str, Any], path: str = ""):
        for key, expected in rule.items():
            required = key.startswith("!")
            clean_key = key.lstrip("!")

            if clean_key not in data:
                if required:
                    errors.append(f"Missing required field: {path}{clean_key}")
                continue

            value = data[clean_key]

            # Validate basic types
            if isinstance(expected, str):
                if expected.lower() in allowed_validation_types:
                    expected_type = expected.strip('<>').lower()
                    if expected_type in type_checks and not type_checks[expected_type](value):
                        errors.append(f"{path}{clean_key} should be a {expected_type}")
                elif expected.startswith("/"):
                    expected_first = expected.find('/')
                    expected_last = expected.rfind('/')
                    if expected_first == -1 or expected_last <= expected_first:
                        errors.append(f"{path}{clean_key} has an invalid regex pattern {expected}")
                    else:
                        pattern = expected[expected_first+1:expected_last]
                        compiled_pattern = get_compiled_pattern(pattern)
                        if not compiled_pattern.fullmatch(str(value)):
                            errors.append(f"{path}{clean_key} does not match the pattern /{pattern}/")
                else:
                    if value != expected:
                        errors.append(f"{path}{clean_key} = {value} should be '{expected}'")

            elif isinstance(expected, dict):
                # Validate nested objects
                if not isinstance(value, dict):
                    errors.append(f"{path}{clean_key} should be an object")
                else:
                    check_structure(value, expected, path + clean_key + ".")

            elif isinstance(expected, list) and expected:
                # Validate lists with expected structure
                if not isinstance(value, list):
                    errors.append(f"{path}{clean_key} should be a list")
                else:
                    for i, item in enumerate(value):
                        check_structure(item, expected[0], f"{path}{clean_key}[{i}].")

    check_structure(event, rules)
    
    return (len(errors) == 0, errors)

def get_output_folder(config, logger):
    """  Get and create output folder if it doesn't exist. """
    # Get output folder from config, default to "." (current directory)
    output_folder = config['config'].get('output_folder', '.')
    
    try:
        # Handle Windows paths with single backslashes by using raw string
        if isinstance(output_folder, str):
            output_folder = rf"{output_folder}"
        
        # Convert to Path object for cross-platform handling
        folder_path = Path(output_folder)
        
        # Convert relative path to absolute
        if not folder_path.is_absolute():
            config_dir = Path(config['_config_file_path']).parent
            folder_path = config_dir / folder_path
        
        # Create folder if it doesn't exist
        folder_path.mkdir(parents=True, exist_ok=True)
        logger.log(f"Output folder confirmed: {folder_path}")
        
    except Exception as e:
        logger.log(f"Error creating output folder: {clean_error_message(e)}", "ERROR")
        # Fall back to script directory
        folder_path = Path().absolute()
        logger.log(f"Using fallback output folder: {folder_path}", "INFO")
    
    return folder_path

def save_results(config, logger, log_data, debug_logs=None):
    """Save results to configured destination"""
    output_destination = config['config'].get('output_destination', 'excel')
    
    if output_destination == 'google_sheets':
        try:
            writer = GoogleSheetsWriter(config, logger)
            return writer.save_data(log_data, debug_logs)
        except Exception as e:
            logger.log(f"Warning: Saving to Google Sheets failed with error: {str(e)}.\nSaving results in Excel file instead", "WARNING")
            writer = ExcelWriter(config, logger)
            return writer.save_data(log_data, debug_logs)
    else:
        writer = ExcelWriter(config, logger)
        return writer.save_data(log_data, debug_logs)

def main(debug_prints=False):
    """Main execution function with optional debug printing"""
    print(PROJECT_HEADER)

    if len(sys.argv) < 2:
        print("Usage: python omdl.py <config_path>")
        sys.exit(1)

    config_path = sys.argv[1]
    if debug_prints:
        print(f"Starting OMDL with configuration: {config_path}")
    
    logger = LogCollector()
    config = load_config(config_path, logger)
    config['_config_file_path'] = config_path
    
    browser = None
    log_data = {}  # Dictionary to store data for each sequence
    
    try:
        if debug_prints:
            logger.log("Initializing OMDL...", "INFO")
        browser = initialize_browser(config, logger)
        
        # Process each sequence
        for sequence_name, sequence in config['sequence'].items():
            if debug_prints:
                logger.log(f"=== Starting sequence: {sequence_name} ===", "INFO")
            
            # Create thread communication objects for this sequence
            event_queue = Queue()
            stop_monitoring = Event()
            
            # Start monitoring thread
            monitored_events = config['config']['track_events']
            monitor_thread = Thread(
                target=start_monitoring_thread,
                args=(browser, monitored_events, event_queue, stop_monitoring, logger, config)
            )
            monitor_thread.daemon = True
            monitor_thread.start()
            
            # Execute step sequence
            sequence_data = perform_sequence(browser, config, event_queue, sequence, logger)
            log_data[sequence_name] = sequence_data  # Store sequence data
            
            # Stop monitoring for this sequence
            stop_monitoring.set()
            monitor_thread.join()
            
        # Save results
        output_path = save_results(config, logger, log_data, 
                                 logger.get_logs() if config['config'].get('debug_mode', False) else None)
        if debug_prints:
            logger.log(f"Results saved to: {output_path}", "INFO")
            
    except Exception as e:
        error_msg = clean_error_message(e)
        if debug_prints:
            logger.log(f"Critical error: {error_msg}", "ERROR")
        # Create error data
        error_data = [
            ['FATAL_ERROR', 'Script Error', 
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
             '', error_msg]
        ]
        log_data = {'Errors': error_data}
        
        try:
            # Try to save error information
            output_path = save_results(config, logger, log_data, 
                                     logger.get_logs() if config['config'].get('debug_mode', False) else None)
            if debug_prints:
                logger.log(f"Error information saved to: {output_path}")
        except Exception as save_error:
            if debug_prints:
                print(f"Could not save error information: {clean_error_message(save_error)}")
            
    finally:
        if browser:
            browser.quit()
        if debug_prints:
            logger.log("\n🎉🎉🎉 Done! 🎉🎉🎉", "INFO")

# start
if __name__ == "__main__":
    main(debug_prints=True)